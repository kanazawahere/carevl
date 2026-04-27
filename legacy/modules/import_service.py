from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from modules import config_loader
from modules import paths
from modules import record_store
from modules import sync


DEFAULT_IMPORT_DIR = Path(paths.get_writable_path("reports/imports"))
DATA_SHEET_NAME = "DuLieu"
GUIDE_SHEET_NAME = "HuongDan"
LOOKUP_SHEET_NAME = "DanhMuc"


def _normalize_text(value: Any) -> str:
    return str(value or "").strip()


def _sanitize_filename_part(value: str) -> str:
    safe = "".join(ch if ch.isalnum() or ch in {"-", "_"} else "-" for ch in _normalize_text(value))
    while "--" in safe:
        safe = safe.replace("--", "-")
    return safe.strip("-_").lower() or "carevl"


def list_packages() -> List[Dict[str, str]]:
    template = config_loader.load_template_form()
    packages: List[Dict[str, str]] = []
    for package in template.get("packages", []):
        package_id = _normalize_text(package.get("id"))
        if not package_id:
            continue
        packages.append(
            {
                "id": package_id,
                "label": _normalize_text(package.get("label")) or package_id,
            }
        )
    return packages


def _package_definition(package_id: str) -> Dict[str, Any]:
    template = config_loader.load_template_form()
    for package in template.get("packages", []):
        if _normalize_text(package.get("id")) == package_id:
            return package
    raise ValueError(f"Không tìm thấy gói khám '{package_id}'.")


def _package_fields(package_id: str) -> List[Dict[str, Any]]:
    package = _package_definition(package_id)
    fields: List[Dict[str, Any]] = []
    for section in package.get("sections", []):
        section_id = _normalize_text(section.get("id"))
        section_label = _normalize_text(section.get("label")) or section_id
        for field in section.get("fields", []):
            field_id = _normalize_text(field.get("id"))
            if not field_id:
                continue
            fields.append(
                {
                    "section_id": section_id,
                    "section_label": section_label,
                    "field_id": field_id,
                    "field_label": _normalize_text(field.get("label")) or field_id,
                    "type": _normalize_text(field.get("type")) or "text",
                    "required": bool(field.get("required")),
                    "options": [str(item).strip() for item in field.get("options", []) if str(item).strip()],
                }
            )
    return fields


def _template_headers(package_id: str) -> List[str]:
    return ["__ghi_chu", "__goi_kham"] + [f"{item['section_id']}.{item['field_id']}" for item in _package_fields(package_id)]


def _field_by_header(package_id: str) -> Dict[str, Dict[str, Any]]:
    return {f"{item['section_id']}.{item['field_id']}": item for item in _package_fields(package_id)}


def _format_excel_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, date):
        return value.strftime("%d-%m-%Y")
    return _normalize_text(value)


def _value_to_cell_text(value: Any, field_type: str) -> str:
    text = _format_excel_date(value)
    if not text:
        return ""
    if field_type == "number":
        try:
            number = float(str(text).replace(",", "."))
        except ValueError:
            raise ValueError(f"giá trị '{text}' không phải số hợp lệ")
        return str(int(number) if number.is_integer() else number)
    if field_type == "date":
        return text
    return text


def get_template_filename(package_id: str, *, extension: str = ".xlsx") -> str:
    station = sync.get_station_info()
    station_id = _sanitize_filename_part(station.get("station_id", "") or "workspace")
    package_safe = _sanitize_filename_part(package_id)
    return f"{station_id}_{package_safe}_template{extension}"


def get_error_report_filename(package_id: str) -> str:
    station = sync.get_station_info()
    station_id = _sanitize_filename_part(station.get("station_id", "") or "workspace")
    package_safe = _sanitize_filename_part(package_id)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return f"{station_id}_{package_safe}_import_errors_{stamp}.txt"


def _guide_lines(package_label: str) -> List[str]:
    return [
        "Mỗi dòng ở sheet DuLieu là 1 lượt khám mới.",
        f"Template hiện tại dành cho gói khám: {package_label}.",
        "Giữ nguyên tên cột header, không xóa cột __goi_kham.",
        "Các ô có dropdown hãy chọn đúng giá trị trong danh sách.",
        "Ngày nhập theo định dạng dd-mm-yyyy. Ví dụ: 24-04-2026.",
        "Có thể mở file bằng Excel rồi lưu lại dưới dạng .xlsx.",
        "Sau khi điền xong, dùng chức năng 'Nhập dữ liệu từ template' trong app để nạp lại.",
    ]


def export_xlsx_template(output_path: str, package_id: str) -> Dict[str, str]:
    DEFAULT_IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    package = _package_definition(package_id)
    package_label = _normalize_text(package.get("label")) or package_id
    fields = _package_fields(package_id)
    headers = _template_headers(package_id)
    field_map = _field_by_header(package_id)

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = DATA_SHEET_NAME
    guide_sheet = workbook.create_sheet(GUIDE_SHEET_NAME)
    lookup_sheet = workbook.create_sheet(LOOKUP_SHEET_NAME)
    lookup_sheet.sheet_state = "hidden"

    header_fill = PatternFill(fill_type="solid", fgColor="DCEBFA")
    required_fill = PatternFill(fill_type="solid", fgColor="FCE8C8")
    hint_fill = PatternFill(fill_type="solid", fgColor="F7FAFD")

    for col_idx, header in enumerate(headers, start=1):
        cell = data_sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

        if header.startswith("__"):
            hint = "Cột hệ thống, giữ nguyên."
        else:
            item = field_map[header]
            requirement = "bắt buộc" if item["required"] else "tùy chọn"
            field_type = item["type"] or "text"
            option_text = f" | chọn: {', '.join(item['options'])}" if item["options"] else ""
            hint = f"{item['section_label']} / {item['field_label']} [{field_type}; {requirement}]{option_text}"
        hint_cell = data_sheet.cell(row=2, column=col_idx, value=hint)
        hint_cell.fill = required_fill if (not header.startswith("__") and field_map[header]["required"]) else hint_fill
        hint_cell.alignment = Alignment(vertical="top", wrap_text=True)

    data_sheet.cell(row=3, column=1, value="")
    data_sheet.cell(row=3, column=2, value=package_id)
    data_sheet.freeze_panes = "A3"
    data_sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}2"

    for col_idx, header in enumerate(headers, start=1):
        width = 20
        if header == "__ghi_chu":
            width = 16
        elif header == "__goi_kham":
            width = 14
        elif header.startswith("conclusion.") or header.startswith("demographics."):
            width = 22
        data_sheet.column_dimensions[get_column_letter(col_idx)].width = width

    data_sheet.row_dimensions[1].height = 24
    data_sheet.row_dimensions[2].height = 42

    lookup_col = 1
    for item in fields:
        if not item["options"]:
            continue
        col_letter = get_column_letter(lookup_col)
        lookup_sheet.cell(row=1, column=lookup_col, value=f"{item['section_id']}.{item['field_id']}")
        for row_idx, option in enumerate(item["options"], start=2):
            lookup_sheet.cell(row=row_idx, column=lookup_col, value=option)
        formula = f"='{LOOKUP_SHEET_NAME}'!${col_letter}$2:${col_letter}${len(item['options']) + 1}"
        target_col = headers.index(f"{item['section_id']}.{item['field_id']}") + 1
        dv = DataValidation(type="list", formula1=formula, allow_blank=not item["required"])
        dv.prompt = f"Chọn giá trị cho {item['field_label']}"
        dv.error = "Giá trị không nằm trong danh sách cho phép."
        data_sheet.add_data_validation(dv)
        dv.add(f"{get_column_letter(target_col)}3:{get_column_letter(target_col)}500")
        lookup_col += 1

    guide_sheet["A1"] = "Hướng dẫn dùng template nhập dữ liệu"
    guide_sheet["A1"].font = Font(size=14, bold=True)
    for idx, line in enumerate(_guide_lines(package_label), start=3):
        guide_sheet.cell(row=idx, column=1, value=f"{idx - 2}. {line}")
    guide_sheet.column_dimensions["A"].width = 110

    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return {
        "path": str(target),
        "message": f"Đã tạo template Excel cho gói {package_label}.",
    }


def export_csv_template(output_path: str, package_id: str) -> Dict[str, str]:
    DEFAULT_IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    package = _package_definition(package_id)
    package_label = _normalize_text(package.get("label")) or package_id
    fields = _package_fields(package_id)

    headers = _template_headers(package_id)
    hint_row = [
        "Xoa dong nay truoc khi import. Ngay theo dinh dang dd-mm-yyyy. Moi dong la 1 luot kham.",
        package_id,
    ]
    for item in fields:
        requirement = "bat buoc" if item["required"] else "tuy chon"
        field_type = item["type"] or "text"
        option_text = f" | chọn: {', '.join(item['options'])}" if item["options"] else ""
        hint_row.append(f"{item['field_label']} [{field_type}; {requirement}]{option_text}")

    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    with open(target, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerow(hint_row)
        writer.writerow(["", package_id, *["" for _ in fields]])

    return {
        "path": str(target),
        "message": f"Đã tạo template CSV cho gói {package_label}.",
    }


def export_template(output_path: str, package_id: str) -> Dict[str, str]:
    suffix = Path(output_path).suffix.lower()
    if suffix == ".csv":
        return export_csv_template(output_path, package_id)
    return export_xlsx_template(output_path, package_id)


def _is_blank_row(row: Dict[str, str]) -> bool:
    for key, value in row.items():
        if key.startswith("__"):
            continue
        if _normalize_text(value):
            return False
    return True


def _ensure_expected_headers(headers: List[str], package_id: str) -> List[str]:
    expected = _template_headers(package_id)
    missing = [header for header in expected if header not in headers]
    return missing


def _iter_csv_rows(source: Path, package_id: str) -> Iterable[Tuple[int, Dict[str, str]]]:
    with open(source, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError("File CSV không có header hợp lệ.")
        missing = _ensure_expected_headers([str(item) for item in reader.fieldnames], package_id)
        if missing:
            raise ValueError(f"File CSV thiếu cột bắt buộc: {', '.join(missing[:8])}")
        for row_index, row in enumerate(reader, start=2):
            yield row_index, {str(key): _normalize_text(value) for key, value in row.items() if key is not None}


def _iter_xlsx_rows(source: Path, package_id: str) -> Iterable[Tuple[int, Dict[str, str]]]:
    workbook = load_workbook(source, data_only=True)
    if DATA_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"File Excel phải có sheet '{DATA_SHEET_NAME}'.")
    sheet = workbook[DATA_SHEET_NAME]
    headers = [_normalize_text(cell.value) for cell in sheet[1]]
    if not any(headers):
        raise ValueError("Sheet dữ liệu không có header hợp lệ.")
    missing = _ensure_expected_headers(headers, package_id)
    if missing:
        raise ValueError(f"File Excel thiếu cột bắt buộc: {', '.join(missing[:8])}")

    for excel_row in range(3, sheet.max_row + 1):
        row: Dict[str, str] = {}
        has_value = False
        for col_idx, header in enumerate(headers, start=1):
            if not header:
                continue
            value = sheet.cell(row=excel_row, column=col_idx).value
            text = _format_excel_date(value)
            if text:
                has_value = True
            row[header] = text
        if not has_value:
            continue
        yield excel_row, row


def _coerce_value(field_type: str, raw_value: Any, field_label: str) -> Any:
    text = _format_excel_date(raw_value)
    if not text:
        return ""
    if field_type == "number":
        try:
            number = float(str(text).replace(",", "."))
        except ValueError:
            raise ValueError(f"{field_label}: giá trị '{text}' không phải số hợp lệ")
        return int(number) if number.is_integer() else number
    if field_type == "date":
        return text
    return text


def _record_from_row(row: Dict[str, str], package_id: str) -> Tuple[Dict[str, Any], str]:
    field_map = _field_by_header(package_id)
    record_data: Dict[str, Any] = {}
    errors: List[str] = []

    for key, item in field_map.items():
        raw_value = row.get(key, "")
        text_value = _normalize_text(raw_value)
        if item["required"] and not text_value:
            errors.append(f"thiếu '{item['field_label']}'")
            continue
        if not text_value:
            continue

        if item["options"] and text_value not in item["options"]:
            errors.append(f"{item['field_label']}: '{text_value}' không nằm trong danh sách cho phép")
            continue

        try:
            value = _coerce_value(item["type"], text_value, item["field_label"])
        except ValueError as exc:
            errors.append(str(exc))
            continue

        section = record_data.setdefault(item["section_id"], {})
        if not isinstance(section, dict):
            section = {}
            record_data[item["section_id"]] = section
        section[item["field_id"]] = value

    conclusion = record_data.get("conclusion", {})
    date_str = _normalize_text(conclusion.get("ngay_kham")) if isinstance(conclusion, dict) else ""
    if not date_str:
        errors.append("thiếu 'Ngày khám'")

    if errors:
        raise ValueError("; ".join(errors))
    return record_data, date_str


def inspect_template(input_path: str, package_id: str) -> Dict[str, Any]:
    source = Path(input_path)
    if not source.exists():
        return {"ok": False, "message": "Không tìm thấy file import đã chọn."}

    try:
        package = _package_definition(package_id)
    except ValueError as exc:
        return {"ok": False, "message": str(exc)}

    suffix = source.suffix.lower()
    if suffix == ".csv":
        row_iter = _iter_csv_rows
        source_type = "CSV"
    elif suffix in {".xlsx", ".xlsm"}:
        row_iter = _iter_xlsx_rows
        source_type = "Excel"
    else:
        return {"ok": False, "message": "Chỉ hỗ trợ xem trước từ file .xlsx, .xlsm hoặc .csv."}

    total_rows = 0
    valid_rows = 0
    skipped_rows = 0
    errors: List[str] = []
    samples: List[Dict[str, str]] = []

    try:
        iterable = row_iter(source, package_id)
        for row_index, row in iterable:
            note_value = _normalize_text((row or {}).get("__ghi_chu", ""))
            if note_value.lower().startswith("xoa dong nay"):
                skipped_rows += 1
                continue

            if _is_blank_row(row):
                skipped_rows += 1
                continue

            total_rows += 1
            row_package_id = _normalize_text(row.get("__goi_kham", "")) or package_id
            if row_package_id != package_id:
                errors.append(f"Dòng {row_index}: gói khám trong file là '{row_package_id}', không khớp với gói đang chọn '{package_id}'.")
                continue

            try:
                record_data, date_str = _record_from_row(row, package_id)
                valid_rows += 1
                demographics = record_data.get("demographics", {}) if isinstance(record_data.get("demographics"), dict) else {}
                conclusion = record_data.get("conclusion", {}) if isinstance(record_data.get("conclusion"), dict) else {}
                if len(samples) < 5:
                    samples.append(
                        {
                            "row": str(row_index),
                            "name": _normalize_text(demographics.get("ho_ten")) or "---",
                            "identity": _normalize_text(demographics.get("ma_dinh_danh")) or "---",
                            "doctor": _normalize_text(conclusion.get("bac_si")) or "---",
                            "date": date_str or "---",
                        }
                    )
            except Exception as exc:
                errors.append(f"Dòng {row_index}: {exc}")
    except Exception as exc:
        return {"ok": False, "message": str(exc)}

    package_label = _normalize_text(package.get("label")) or package_id
    message = f"Đã quét file {source_type} cho gói {package_label}."
    if total_rows == 0 and not errors:
        message = "File hiện chưa có dòng dữ liệu nào để nhập."

    return {
        "ok": True,
        "message": message,
        "path": str(source),
        "source_type": source_type,
        "total_rows": total_rows,
        "valid_rows": valid_rows,
        "skipped_rows": skipped_rows,
        "error_count": len(errors),
        "errors": errors,
        "samples": samples,
    }


def import_template(input_path: str, package_id: str, author: str) -> Dict[str, Any]:
    source = Path(input_path)
    if not source.exists():
        return {"ok": False, "message": "Không tìm thấy file import đã chọn."}

    try:
        package = _package_definition(package_id)
    except ValueError as exc:
        return {"ok": False, "message": str(exc)}

    suffix = source.suffix.lower()
    if suffix == ".csv":
        row_iter = _iter_csv_rows
        source_type = "CSV"
    elif suffix in {".xlsx", ".xlsm"}:
        row_iter = _iter_xlsx_rows
        source_type = "Excel"
    else:
        return {"ok": False, "message": "Chỉ hỗ trợ import từ file .xlsx, .xlsm hoặc .csv."}

    created = 0
    skipped = 0
    errors: List[str] = []

    try:
        iterable = row_iter(source, package_id)
        for row_index, row in iterable:
            note_value = _normalize_text((row or {}).get("__ghi_chu", ""))
            if note_value.lower().startswith("xoa dong nay"):
                skipped += 1
                continue

            if _is_blank_row(row):
                skipped += 1
                continue

            row_package_id = _normalize_text(row.get("__goi_kham", "")) or package_id
            if row_package_id != package_id:
                errors.append(f"Dòng {row_index}: gói khám trong file là '{row_package_id}', không khớp với gói đang chọn '{package_id}'.")
                continue

            try:
                record_data, date_str = _record_from_row(row, package_id)
                record_store.create(record_data, package_id, author, date_str)
                created += 1
            except Exception as exc:
                errors.append(f"Dòng {row_index}: {exc}")
    except Exception as exc:
        return {"ok": False, "message": str(exc), "created": 0, "skipped": 0, "errors": []}

    package_label = _normalize_text(package.get("label")) or package_id
    if created == 0 and errors:
        return {
            "ok": False,
            "message": f"Không import được dòng dữ liệu nào từ file {source_type} của gói {package_label}.",
            "created": created,
            "skipped": skipped,
            "errors": errors,
        }

    return {
        "ok": True,
        "message": f"Đã nhập {created} lượt khám từ template {source_type} cho gói {package_label}.",
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "path": str(source),
    }


def write_import_error_report(result: Dict[str, Any], package_id: str, output_path: str | None = None) -> Dict[str, str]:
    DEFAULT_IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    target = Path(output_path) if output_path else DEFAULT_IMPORT_DIR / get_error_report_filename(package_id)
    target.parent.mkdir(parents=True, exist_ok=True)

    lines = [
        "Bao cao loi import CareVL",
        f"Thoi gian: {datetime.now().strftime('%H:%M:%S %d-%m-%Y')}",
        f"Goi kham: {package_id}",
        f"Thong diep: {result.get('message', '')}",
        f"Tao moi: {result.get('created', 0)}",
        f"Bo qua: {result.get('skipped', 0)}",
        f"File nguon: {result.get('path', '')}",
        "",
        "Chi tiet loi:",
    ]

    errors = result.get("errors", [])
    if isinstance(errors, list) and errors:
        lines.extend(str(item) for item in errors)
    else:
        lines.append("Khong co loi chi tiet.")

    target.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")
    return {
        "path": str(target),
        "message": "Đã xuất báo cáo lỗi import.",
    }
